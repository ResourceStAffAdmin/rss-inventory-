from __future__ import annotations

import re
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "docs" / "Inventory - (1).xlsx"
OUTPUT = ROOT / "database" / "seeders" / "004_import_inventory_workbook.sql"
BROAD_CATEGORIES = {
    "Computers": "Laptops, desktops, and computer units",
    "Displays": "Monitors and display equipment",
    "Peripherals": "Keyboards, mice, headsets, webcams, and input devices",
    "Cables & Adapters": "Cables, converters, display adapters, and connectors",
    "Power & Network": "UPS units, power cables, and network adapters",
    "Accessories": "Bags and other supporting accessories",
}


def clean(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def sql(value: object) -> str:
    if value is None:
        return "NULL"
    text = str(value).replace("\\", "\\\\").replace("'", "''")
    return f"'{text}'"


def generated_sku(name: str, row_number: int) -> str:
    base = re.sub(r"[^A-Za-z0-9]+", "-", name.upper()).strip("-") or "ITEM"
    return f"{base[:45]}-{row_number:04d}"[:60]


def category_for(name: str) -> str:
    upper = name.upper()
    if "MONITOR" in upper:
        return "Displays"
    if upper in {"LAPTOP", "PC"} or "LAPTOP MAC" in upper:
        return "Computers"
    if any(token in upper for token in ("MOUSE", "KEYBOARD", "HEADSET", "WEBCAM")):
        return "Peripherals"
    if any(token in upper for token in ("UPS", "POWER", "WIFI", "NETWORK")):
        return "Power & Network"
    if any(token in upper for token in ("HDMI", "USB", "DP", "ADAPTER", "CONVERTER")):
        return "Cables & Adapters"
    return "Accessories"


def load_records() -> list[dict[str, object]]:
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb["STOCKS"]

    rows = list(ws.iter_rows(values_only=True))
    header = next(
        i
        for i, row in enumerate(rows)
        if {"ID", "Category", "Available Stocks"}.issubset({clean(v) for v in row})
    )
    headers = [clean(v) for v in rows[header]]
    id_col = headers.index("ID")
    name_col = headers.index("Category")
    stock_col = headers.index("Available Stocks")

    records: list[dict[str, object]] = []
    for index, row in enumerate(rows[header + 1 :], start=header + 2):
        item_id = clean(row[id_col] if id_col < len(row) else "")
        name = clean(row[name_col] if name_col < len(row) else "")
        if not item_id or not name:
            continue
        quantity = row[stock_col] if stock_col < len(row) else 0
        quantity = float(quantity or 0)
        records.append(
            {
                "sku": item_id[:60],
                "name": name[:191],
                "category": category_for(name),
                "description": None,
                "quantity": quantity,
            }
        )
    return records


def load_legacy_detail_skus() -> list[str]:
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb["STOCKS"]
    rows = list(ws.iter_rows(values_only=True))
    header = next(i for i, row in enumerate(rows) if "ITEM" in [clean(v).upper() for v in row])
    item_col = [clean(v).upper() for v in rows[header]].index("ITEM")

    skus: list[str] = []
    for index, row in enumerate(rows[header + 1 :], start=header + 2):
        name = clean(row[item_col] if item_col < len(row) else "")
        if name:
            skus.append(generated_sku(name, index))
    return sorted(set(skus))


def main() -> None:
    records = load_records()
    legacy_skus = load_legacy_detail_skus()
    lines = [
        "SET NAMES utf8mb4;",
        "START TRANSACTION;",
        "",
        "SET @location_id := (SELECT id FROM locations ORDER BY id ASC LIMIT 1);",
        "SET @moved_by := (SELECT id FROM users ORDER BY id ASC LIMIT 1);",
        "",
        "UPDATE categories",
        f"SET is_active = 0 WHERE name NOT IN ({', '.join(sql(name) for name in BROAD_CATEGORIES)});",
        "",
    ]

    for category, description in BROAD_CATEGORIES.items():
        lines.extend(
            [
                f"INSERT INTO categories (name, description, is_active) VALUES ({sql(category)}, {sql(description)}, 1)",
                "ON DUPLICATE KEY UPDATE description = VALUES(description), is_active = 1;",
                "",
            ]
        )

    if legacy_skus:
        lines.extend(
            [
                "UPDATE products",
                f"SET is_active = 0 WHERE sku IN ({', '.join(sql(sku) for sku in legacy_skus)});",
                "",
            ]
        )

    for record in records:
        lines.extend(
            [
                f"SET @category_id := (SELECT id FROM categories WHERE name = {sql(record['category'])} LIMIT 1);",
                f"SET @sku := {sql(record['sku'])};",
                "SET @product_id := (SELECT id FROM products WHERE sku = @sku LIMIT 1);",
                "SET @previous_stock := IF(@product_id IS NULL, 0, COALESCE((SELECT quantity_on_hand FROM inventory_balances WHERE location_id = @location_id AND product_id = @product_id), 0));",
                "INSERT INTO products (sku, name, description, category_id, preferred_supplier_id, unit_of_measure, cost_price, sell_price, reorder_level, is_active)",
                f"VALUES (@sku, {sql(record['name'])}, {sql(record['description'])}, @category_id, NULL, 'pcs', 0, 0, 0, 1)",
                "ON DUPLICATE KEY UPDATE",
                f"    name = {sql(record['name'])},",
                f"    description = {sql(record['description'])},",
                "    category_id = @category_id,",
                "    unit_of_measure = 'pcs',",
                "    reorder_level = 0,",
                "    is_active = 1;",
                "SET @product_id := (SELECT id FROM products WHERE sku = @sku LIMIT 1);",
                "INSERT INTO inventory_balances (location_id, product_id, quantity_on_hand, quantity_reserved)",
                f"VALUES (@location_id, @product_id, {record['quantity']:.3f}, 0)",
                f"ON DUPLICATE KEY UPDATE quantity_on_hand = {record['quantity']:.3f}, updated_at = CURRENT_TIMESTAMP;",
                "INSERT INTO inventory_movements (movement_type, product_id, source_location_id, destination_location_id, quantity, unit_cost, reference_type, reference_id, reason, moved_by, moved_at, previous_stock, new_stock, status)",
                f"SELECT 'ADJUSTMENT', @product_id, NULL, @location_id, ABS({record['quantity']:.3f} - @previous_stock), NULL, 'MANUAL', NULL, 'Imported from Inventory - (1).xlsx', @moved_by, NOW(), @previous_stock, {record['quantity']:.3f}, 'COMPLETED'",
                f"WHERE ABS({record['quantity']:.3f} - @previous_stock) > 0.0001;",
                "",
            ]
        )

    lines.extend(["COMMIT;", ""])
    OUTPUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {OUTPUT} with {len(records)} records")


if __name__ == "__main__":
    main()
